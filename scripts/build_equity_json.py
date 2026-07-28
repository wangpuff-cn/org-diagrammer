#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
org-diagrammer · build_equity_json.py
解析股权/组织架构 Excel 模版 → 计算布局 → 输出前端 JSON（或注入模板直接产出独立 HTML）。

用法:
  python3 build_equity_json.py 输入.xlsx [--mode auto|equity|org] [--title 标题]
          [--json out.json] [--html out.html] [--template 模板路径]

--html 缺省模板: <script_dir>/../assets/widget-template/index.html
JSON 契约(供 assets/widget-template/index.html 使用):
  { viewBox:{x,y,w,h},
    nodes:[{id,name,geo,type,x,y,w,h,nameFont,layer, face?,text?,
            function?,leader?,headcount?,regular?,outsourced?}],
    edges:[{from,to,ratio?,style}],          # style: solid|dashed
    wires:[{pts:[[x,y]..],style,arrow,from,to}],
    labels:[{x,y,text,from,to}] }
"""
import argparse, colorsys, json, re, sys, zipfile
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

# ---------------- 视觉参数（随需可调，机制勿动） ----------------
H_UNIT        = 42.0          # 节点高度
NAME_FONT_MAX = 14            # 名称行最大字号
NAME_FONT_MIN = 6             # 名称行最小字号
GEO_FONT      = 12            # 地区行字号
TEXT_COEFF    = 1.15          # 字号→宽度换算系数（紧凑贴合）
SAFETY        = 0.95          # 可用宽度安全系数
RATIOS        = [3, 4, 5, 6, 7, 8, 9]      # 宽度档 = k * H_UNIT
NODE_PAD      = H_UNIT * 0.15              # 名称两侧合计留白（贴合文字）
NODE_GAP      = H_UNIT * 0.6               # 同层相邻节点间距 = 25.2（与前端 X_GAP 一致）
GAP_SIMPLE    = H_UNIT * 1.0               # 一对一上下层空白
GAP_MULTI     = H_UNIT * 2.0               # fan-in/fan-out 上下层空白
WRAP_ROW_GAP  = H_UNIT * 0.6               # 同逻辑层相邻视觉行空白
MAX_PER_ROW_EQUITY = 10        # 股权图: 同一母公司同层超过即换行
MAX_PER_ROW_ORG    = 15        # 组织图: 阈值放宽


# ---------------- 宽度/字号估算 ----------------
def char_width(s):
    w = 0.0
    for c in str(s):
        if '\u4e00' <= c <= '\u9fff' or c in '（）【】·':
            w += 1.0
        elif c.isascii() and c.isalpha():
            w += 0.6
        else:
            w += 0.8
    return w

def fit_fontsize(name, node_w):
    avail = node_w * SAFETY - NODE_PAD
    cw = char_width(name)
    if cw <= 0:
        return NAME_FONT_MAX
    return max(NAME_FONT_MIN, min(NAME_FONT_MAX, avail / (cw * TEXT_COEFF)))

def raw_need_w(node):
    nw = char_width(node['name']) * NAME_FONT_MAX * TEXT_COEFF
    gw = char_width(node.get('geo', '')) * GEO_FONT * TEXT_COEFF if node.get('geo') else 0
    return max(nw, gw) + NODE_PAD

def pick_slot(raw_w):
    # 不再按 RATIOS 整数倍取档（量化进位会造成大量留白）；
    # 直接按文字实际需求宽度贴合，仅保留最小宽度下限，避免过窄。
    return max(raw_w, H_UNIT * 2)

def unify_widths(nodes, mode):
    """同一层级节点等宽: 以该层最长名称节点的需求宽度为准, 应用到全层; 宽度贴合文字。"""
    key = lambda n: n['layer']
    best = {}
    for n in nodes:
        k = key(n)
        best[k] = max(best.get(k, 0.0), pick_slot(raw_need_w(n)))
    for n in nodes:
        n['w'] = round(best[key(n)], 1)
        n['h'] = H_UNIT
        n['nameFont'] = round(fit_fontsize(n['name'], n['w']), 1)


# ---------------- Excel 填充色（主题色 + tint 还原） ----------------
def load_theme_colors(xlsx_path):
    try:
        with zipfile.ZipFile(xlsx_path) as z:
            xml = z.read('xl/theme/theme1.xml').decode('utf-8')
    except Exception:
        return ['FFFFFF'] * 12
    seg = re.search(r'<a:clrScheme.*?</a:clrScheme>', xml, re.S)
    if not seg:
        return ['FFFFFF'] * 12
    seg = seg.group(0)
    order = ['dk1', 'lt1', 'dk2', 'lt2', 'accent1', 'accent2', 'accent3',
             'accent4', 'accent5', 'accent6', 'hlink', 'folHlink']
    got = {}
    for nm, body in re.findall(r'<a:(\w+)>(.*?)</a:\1>', seg, re.S):
        m = re.search(r'srgbClr val="([0-9A-Fa-f]{6})"', body) or \
            re.search(r'sysClr[^>]*lastClr="([0-9A-Fa-f]{6})"', body)
        if m and nm in order:
            got[nm] = m.group(1).upper()
    return [got.get(n, 'FFFFFF') for n in order]

def apply_tint(hex6, tint):
    r, g, b = (int(hex6[i:i+2], 16) / 255 for i in (0, 2, 4))
    h, l, s = colorsys.rgb_to_hls(r, g, b)
    l = l * (1.0 + tint) if tint < 0 else (l * (1.0 - tint) + tint if tint > 0 else l)
    r, g, b = colorsys.hls_to_rgb(h, max(0, min(1, l)), s)
    return '#%02X%02X%02X' % (round(r*255), round(g*255), round(b*255))

def cell_fill_hex(cell, theme):
    fill = cell.fill
    if not fill or fill.patternType != 'solid':
        return None
    fg = fill.fgColor
    try:
        rgb = fg.rgb
        if isinstance(rgb, str) and re.fullmatch(r'[0-9A-Fa-f]{8}', rgb):
            h = '#' + rgb[2:].upper()
            return None if h == '#000000' and fg.type != 'rgb' else h
    except Exception:
        pass
    if fg.type == 'theme' and fg.theme is not None and fg.theme < len(theme):
        h = apply_tint(theme[fg.theme], fg.tint or 0.0)
        return None if h.upper() in ('#FFFFFF', '#000000') and not (fg.tint or 0) else h
    return None

def text_color_for(bg):
    if not bg:
        return '#1a1a1a'
    r, g, b = (int(bg[1+i:3+i], 16) / 255 for i in (0, 2, 4))
    return '#FFFFFF' if (0.299*r + 0.587*g + 0.114*b) < 0.55 else '#1a1a1a'


# ---------------- 通用取值 ----------------
def sval(v):
    s = ('' if v is None else str(v)).strip()
    return '' if s in ('-', '——', '—', 'None') else s

def geo_str(*parts):
    return ' · '.join(p for p in (sval(x) for x in parts) if p)

def fmt_ratio(v):
    s = sval(v)
    if not s:
        return ''
    if s.endswith('%'):
        return s
    try:
        f = float(s)
        f = f * 100 if f <= 1 else f
        return f'{f:g}%'
    except ValueError:
        return s

COMPANY_KW = (r'公司|企业|合伙|中心|集团|基金|厂|店|院|部|事务所|'
              r'Limited|Ltd\.?|Inc\.?|Co\.?,?|Company|Corp\.?|Corporation|Group|Holdings?|'
              r'LLC|LLP|\bLP\b|GmbH|S\.A\.|Pte|Pty|B\.V\.')
OFFSHORE_KW = r'开曼|Cayman|BVI|英属维京|维尔京|香港|Hong\s*Kong|\bHK\b'

def classify(name):
    """公司类关键词 → offshore/domestic；无公司关键词的短名称(中英文人名) → 自然人。
    注意: 境内/境外列的显式标注优先级更高（load_equity 中二次校正）。"""
    n = str(name)
    if re.search(r'上市|NYSE|NASDAQ|HKEX|交易所|股票代码', n):
        return 'listed'
    if re.search(COMPANY_KW, n, re.I):
        return 'offshore' if re.search(OFFSHORE_KW, n, re.I) else 'domestic'
    if re.fullmatch(r'[\u4e00-\u9fff]{2,4}', n) or re.fullmatch(r"[A-Za-z\.\-]+( [A-Za-z\.\-]+){0,2}", n):
        return 'natural_person'
    return 'domestic'


# ---------------- 解析：股权模版 ----------------
def load_equity(path):
    theme = load_theme_colors(path)
    wb  = openpyxl.load_workbook(path)            # 保留样式（读 I 列填充色）
    wbv = openpyxl.load_workbook(path, data_only=True)
    sheet = next((s for s in wb.sheetnames if '股权' in s), wb.sheetnames[0])
    ws, wsv = wb[sheet], wbv[sheet]
    H = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    def col(*keys):
        for k, v in H.items():
            if k and any(kk in str(k) for kk in keys):
                return v
        return None
    cL, cN = col('层级', 'Level'), col('公司名称', 'Company')
    cC, cP, cCity, cD = col('国家', 'Country'), col('省', 'Province'), col('市', 'City'), col('区', 'District')
    cR, cDom, cColor = col('股权关系', 'Relationship'), col('境内', 'Domestic'), col('节点颜色', 'Color')
    cHolder, cPct, cNote = col('股东', 'Shareholder'), col('持股比例', 'Ownership'), col('特殊要求', 'Notes')
    cBiz = col('主体业务', '业务及职能', '职能', 'Business', 'Function')   # 主体业务及职能（悬停展示）

    nodes, edges, notes = {}, [], []
    for r in range(2, ws.max_row + 1):
        name = sval(ws.cell(r, cN).value if cN else None)
        if not name:
            continue
        face = cell_fill_hex(ws.cell(r, cColor), theme) if cColor else None
        dom = sval(ws.cell(r, cDom).value if cDom else '')
        note = sval(ws.cell(r, cNote).value if cNote else '')
        if note:
            notes.append(note)
        biz = sval(wsv.cell(r, cBiz).value if cBiz else '')
        node = {'id': name, 'name': name,
                'geo': geo_str(ws.cell(r, cC).value if cC else '',
                               ws.cell(r, cP).value if cP else '',
                               ws.cell(r, cCity).value if cCity else '',
                               ws.cell(r, cD).value if cD else ''),
                'type': classify(name),
                'layer': int(ws.cell(r, cL).value or 0),
                'domestic': ('境内' in dom or 'domestic' in dom.lower())}
        if biz:
            node['function'] = biz                 # 悬停展示「主体业务及职能」
            if '上市' in biz or 'listed' in biz.lower():
                node['type'] = 'listed'            # 业务职能标注上市 → 上市主体（加粗边框）
        # 境内/境外列显式标注优先于名称推断（如 WFOE 名称含 Ltd. 但属境内）
        if node['type'] not in ('listed', 'natural_person'):
            node['type'] = 'domestic' if node['domestic'] else 'offshore'
        if face:
            node['face'] = face
            node['text'] = text_color_for(face)
        nodes[name] = node
        rel = sval(ws.cell(r, cR).value if cR else '')
        holder_raw = wsv.cell(r, cHolder).value if cHolder else None
        if holder_raw is None and cHolder:
            holder_raw = ws.cell(r, cHolder).value
        holder = sval(holder_raw)
        pct = wsv.cell(r, cPct).value if cPct else None
        if holder and '顶层' not in holder and 'top' not in holder.lower() and rel:
            style = 'dashed' if ('协议控制' in rel or 'vie' in rel.lower() or 'contractual' in rel.lower()) else 'solid'
            holders = [h.strip() for h in holder.split('/') if h.strip()]
            pcts = [p.strip() for p in str(pct).split('/')] if pct is not None and '/' in str(pct) else [pct]
            for i, h in enumerate(holders):
                raw = pcts[i] if i < len(pcts) else (pcts[0] if len(pcts) == 1 else '')
                edges.append({'from': h, 'to': name, 'ratio': fmt_ratio(raw), 'style': style})
    return list(nodes.values()), edges, notes


# ---------------- 解析：组织模版 ----------------
def load_org(path):
    theme = load_theme_colors(path)
    wb  = openpyxl.load_workbook(path)
    wbv = openpyxl.load_workbook(path, data_only=True)
    sheet = next((s for s in wb.sheetnames if '组织' in s or '部门' in s), wb.sheetnames[0])
    ws, wsv = wb[sheet], wbv[sheet]
    H = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    def col(*keys):
        for k, v in H.items():
            if k and any(kk in str(k) for kk in keys):
                return v
        return None
    cColor = col('节点颜色', 'Color')
    cL = col('层级', 'Level')
    cN = col('部门', '组织名称', 'Department', 'Organization')
    cLeader = col('负责人', 'Leader')
    cFunc = col('职能', 'Function')
    cHead = col('人数', 'Headcount')
    cRegular = col('正编', 'Regular')
    cOut = col('外包', 'Outsourced')
    cRel = col('管理关系', '股权关系', 'Management', 'Relationship')
    cUp = col('上级', 'Parent')
    cNote = col('特殊要求', 'Notes')

    nodes, edges, notes = {}, [], []
    for r in range(2, ws.max_row + 1):
        name = sval(ws.cell(r, cN).value if cN else None)
        if not name:
            continue
        face = cell_fill_hex(ws.cell(r, cColor), theme) if cColor else None
        note = sval(ws.cell(r, cNote).value if cNote else '')
        if note:
            notes.append(note)
        node = {'id': name, 'name': name, 'geo': '', 'type': 'org',
                'layer': int(ws.cell(r, cL).value or 0)}
        leader = sval(wsv.cell(r, cLeader).value if cLeader else '')
        func = sval(wsv.cell(r, cFunc).value if cFunc else '')
        if leader:
            node['leader'] = leader
        if func:
            node['function'] = func
        if cHead is not None:
            hv = wsv.cell(r, cHead).value
            if isinstance(hv, (int, float)):
                node['headcount'] = int(hv)
            elif hv is not None:
                # 文本形如 "正编 156 外包 69"：拆分正则提取；纯数字则作为总人数
                m1 = re.search(r'正编\s*(\d+)', str(hv))
                m2 = re.search(r'外包\s*(\d+)', str(hv))
                if m1:
                    node['regular'] = int(m1.group(1))
                if m2:
                    node['outsourced'] = int(m2.group(1))
                if not m1 and not m2:
                    m0 = re.search(r'(\d+)', str(hv))
                    if m0:
                        node['headcount'] = int(m0.group(1))
        else:
            for key, cc in (('regular', cRegular), ('outsourced', cOut)):
                if cc is not None:
                    vv = wsv.cell(r, cc).value
                    if isinstance(vv, (int, float)):
                        node[key] = int(vv)
        if face:
            node['face'] = face
            node['text'] = text_color_for(face)
        nodes[name] = node
        rel = sval(ws.cell(r, cRel).value if cRel else '')
        # 上级列可能是公式（如 =C$2），须读 data_only 缓存值；无缓存时回退公式版
        up_raw = wsv.cell(r, cUp).value if cUp else None
        if up_raw is None and cUp:
            up_raw = ws.cell(r, cUp).value
        up = sval(up_raw)
        if up and '顶层' not in up and 'top' not in up.lower() and rel:
            style = 'dashed' if re.search(r'虚线|兼管|协议|dashed|virtual|vie|contractual', rel, re.I) else 'solid'
            for h in [x.strip() for x in up.split('/') if x.strip()]:
                edges.append({'from': h, 'to': name, 'ratio': '', 'style': style})
    return list(nodes.values()), edges, notes


# ---------------- L 列特殊要求 → 显示层 dlayer ----------------
def apply_special_rules(nodes, notes):
    for n in nodes:
        n['dlayer'] = n['layer']
    if any(('境内' in s and '境外' in s and ('低于' in s or '下' in s)) or
           ('domestic' in s.lower() and 'offshore' in s.lower() and 'below' in s.lower())
           for s in notes):
        off = [n['layer'] for n in nodes if not n.get('domestic')]
        dom = [n for n in nodes if n.get('domestic')]
        if off and dom:
            offset = max(off) + 1 - min(n['layer'] for n in dom)
            if offset > 0:
                for n in dom:
                    n['dlayer'] = n['layer'] + offset


# ---------------- 布局 ----------------
def layer_gap(prev_ids, cur_ids, edges):
    rel = [e for e in edges if e['from'] in prev_ids and e['to'] in cur_ids]
    od = Counter(e['from'] for e in rel)
    idg = Counter(e['to'] for e in rel)
    fan = any(v > 1 for v in od.values()) or any(v > 1 for v in idg.values())
    return GAP_MULTI if fan else GAP_SIMPLE

def layout(nodes, edges, max_per_row):
    nb = {n['id']: n for n in nodes}
    children_of, parents_of = defaultdict(list), defaultdict(list)
    for e in edges:
        if e['from'] in nb and e['to'] in nb:
            children_of[e['from']].append(e['to'])
            parents_of[e['to']].append(e['from'])
    DL = lambda n: n.get('dlayer', n['layer'])
    layers = defaultdict(list)
    for n in nodes:
        layers[DL(n)].append(n)

    # (1) tidy-tree 基础 x：叶子从左到右，父居中于子
    has_parent = set(parents_of)
    roots = [n['id'] for n in nodes if n['id'] not in has_parent]
    base_x, visited, cursor = {}, set(), [0.0]
    def place(nid):
        """子树占位宽度 = max(自身宽度, 子代跨度)。
        父比子树宽时：子代块整体右移居中于父，游标按父宽前进——
        否则窄子树的宽父会与右侧兄弟子树间距被吃掉甚至重叠。"""
        if nid in visited:
            return base_x[nid]
        visited.add(nid)
        w = nb[nid]['w']
        kids = [k for k in children_of.get(nid, []) if k not in visited]
        if not kids:
            base_x[nid] = cursor[0] + w / 2
            cursor[0] += w + NODE_GAP
            return base_x[nid]
        start = cursor[0]
        kxs = [place(k) for k in kids]
        span = cursor[0] - start - NODE_GAP          # 子代块实际跨度
        if w > span:
            dx = (w - span) / 2
            def shift(c):
                base_x[c] += dx
                for cc in children_of.get(c, []):
                    if cc in base_x:
                        shift(cc)
            for k in kids:
                shift(k)
            base_x[nid] = start + w / 2
            cursor[0] = start + w + NODE_GAP
        else:
            base_x[nid] = (min(kxs) + max(kxs)) / 2
        return base_x[nid]
    for r in roots:
        place(r)
    for nid in nb:
        if nid not in base_x:
            place(nid)

    # (2) 兄弟组（同一直接上级 + 同一显示层）超限换行；有下层的放靠下行
    def sibling_key(n):
        ps = [p for p in parents_of.get(n['id'], []) if DL(nb[p]) < DL(n)]
        return tuple(sorted(ps)) if ps else ('__ROOT__',)
    def split_rows(sibs):
        ordered = sorted(sibs, key=lambda n: base_x[n['id']])
        if len(ordered) <= max_per_row:
            return [ordered]
        leaves, branches = [], []
        for n in ordered:
            has_lower = any(DL(nb[c]) > DL(n) for c in children_of.get(n['id'], []))
            (branches if has_lower else leaves).append(n)
        rc = (len(ordered) + max_per_row - 1) // max_per_row
        rows = [[] for _ in range(rc)]
        remain = list(branches)
        for ri in range(rc - 1, -1, -1):
            take = min(max_per_row, len(remain))
            if take:
                rows[ri] = remain[-take:]
                remain = remain[:-take]
        pos = 0
        for ri in range(rc):
            room = max_per_row - len(rows[ri])
            take = min(room, len(leaves) - pos)
            if take:
                rows[ri] = leaves[pos:pos+take] + rows[ri]
                pos += take
        return rows

    family_rows = {}
    row_count = {}
    for lv in sorted(layers):
        fams = defaultdict(list)
        for n in layers[lv]:
            fams[sibling_key(n)].append(n)
        family_rows[lv] = {k: split_rows(v) for k, v in fams.items()}
        row_count[lv] = max(len(v) for v in family_rows[lv].values())

    # (3) 最终坐标（y 向下递增）
    row_y = {}
    prev_bottom, prev_ids = None, set()
    for i, lv in enumerate(sorted(layers)):
        cur_ids = {n['id'] for n in layers[lv]}
        top = H_UNIT if i == 0 else prev_bottom + layer_gap(prev_ids, cur_ids, edges)
        for ri in range(row_count[lv]):
            row_y[(lv, ri)] = top + H_UNIT / 2 + ri * (H_UNIT + WRAP_ROW_GAP)
        for rows in family_rows[lv].values():
            sibs = [n for row in rows for n in row]
            if len(sibs) <= max_per_row:
                for n in rows[0]:
                    n['wrap_row'] = 0
                    n['x'] = round(base_x[n['id']], 1)
                    n['y'] = round(row_y[(lv, 0)], 1)
                continue
            cc = min(max_per_row, len(sibs))
            cw = max(n['w'] for n in sibs)
            grid_w = cc * cw + (cc - 1) * NODE_GAP
            ctr = (min(base_x[n['id']] for n in sibs) + max(base_x[n['id']] for n in sibs)) / 2
            left = ctr - grid_w / 2
            for ri, row in enumerate(rows):
                for ci, n in enumerate(row):
                    n['wrap_row'] = ri
                    n['x'] = round(left + cw / 2 + ci * (cw + NODE_GAP), 1)
                    n['y'] = round(row_y[(lv, ri)], 1)
        prev_bottom = top + row_count[lv] * H_UNIT + (row_count[lv] - 1) * WRAP_ROW_GAP
        prev_ids = cur_ids


# ---------------- 连线 + 标注（与前端 reroute 同逻辑；干线在上下层间居中） ----------------
def build_wires(nodes, edges):
    nb = {n['id']: n for n in nodes}
    edges = [e for e in edges if e['from'] in nb and e['to'] in nb]
    top = lambda nid: (nb[nid]['x'], nb[nid]['y'] - nb[nid]['h'] / 2)
    bot = lambda nid: (nb[nid]['x'], nb[nid]['y'] + nb[nid]['h'] / 2)
    wires, labels, consumed = [], [], set()
    def wire(pts, style, arrow, frm, to):
        wires.append({'pts': [[round(x, 1), round(y, 1)] for x, y in pts],
                      'style': style, 'arrow': arrow, 'from': frm, 'to': to})
    def label(x, y, text, frm, to):
        if text:
            labels.append({'x': round(x + 4, 1), 'y': round(y, 1),
                           'text': text, 'from': frm, 'to': to})

    by_child, by_parent = defaultdict(list), defaultdict(list)
    for e in edges:
        by_child[e['to']].append(e)
        by_parent[e['from']].append(e)

    # fan-in：一子多父，父同排共用汇线
    for to, es in by_child.items():
        if len(es) < 2:
            continue
        groups = defaultdict(list)
        for e in es:
            groups[nb[e['from']]['y']].append(e)
        for g in groups.values():
            if len(g) < 2:
                continue
            cx, cy = top(to)
            ybus = (bot(g[0]['from'])[1] + cy) / 2
            xs = []
            for e in g:
                bx, by = bot(e['from'])
                xs.append(bx)
                wire([(bx, by), (bx, ybus)], e['style'], False, e['from'], to)
                label(bx, (by + ybus) / 2, e.get('ratio'), e['from'], to)
                consumed.add(id(e))
            wire([(min(xs + [cx]), ybus), (max(xs + [cx]), ybus)], g[0]['style'], False, '\x01'.join(e['from'] for e in g), to)
            wire([(cx, ybus), (cx, cy)], g[0]['style'], True, g[0]['from'], to)

    # fan-out / 一对一
    for frm, es0 in by_parent.items():
        es = [e for e in es0 if id(e) not in consumed]
        if not es:
            continue
        groups = defaultdict(list)
        for e in es:
            groups[nb[e['to']]['y']].append(e)
        for g in groups.values():
            px, py = bot(frm)
            if len(g) == 1:
                e = g[0]
                cx, cy = top(e['to'])
                ybus = (py + cy) / 2
                if abs(px - cx) < 1:
                    wire([(px, py), (cx, cy)], e['style'], True, frm, e['to'])
                    label(px, (py + cy) / 2, e.get('ratio'), frm, e['to'])
                else:
                    wire([(px, py), (px, ybus)], e['style'], False, frm, e['to'])
                    wire([(px, ybus), (cx, ybus)], e['style'], False, frm, e['to'])
                    wire([(cx, ybus), (cx, cy)], e['style'], True, frm, e['to'])
                    label(cx, (ybus + cy) / 2, e.get('ratio'), frm, e['to'])
            else:
                cs = [top(e['to']) for e in g]
                ybus = (py + cs[0][1]) / 2
                # 不画通长干线：父竖线归属全部子节点；水平段按子节点各自绘制(父正下方→子正上方)，
                # 重叠部分视觉等同一条干线，但高亮只亮链路那一段，不再溢出
                all_to = '\x01'.join(e['to'] for e in g)
                wire([(px, py), (px, ybus)], g[0]['style'], False, frm, all_to)
                for e, (cx, cy) in zip(g, cs):
                    wire([(min(px, cx), ybus), (max(px, cx), ybus)], e['style'], False, frm, e['to'])
                    wire([(cx, ybus), (cx, cy)], e['style'], True, frm, e['to'])
                    label(cx, (ybus + cy) / 2, e.get('ratio'), frm, e['to'])
    return wires, labels


# ---------------- 主流程 ----------------
def detect_mode(path):
    wb = openpyxl.load_workbook(path, read_only=True)
    for s in wb.sheetnames:
        sl = s.lower()
        if '组织' in s or '部门' in s or 'department' in sl or 'org' in sl:
            return 'org'
    ws = wb[wb.sheetnames[0]]
    heads = [str(ws.cell(1, c).value or '') for c in range(1, ws.max_column + 1)]
    wb.close()
    return 'org' if any('部门' in h or '组织名称' in h or 'department' in h.lower()
                      or 'organization' in h.lower() for h in heads) else 'equity'

def build(path, mode):
    if mode == 'org':
        nodes, edges, notes = load_org(path)
    else:
        nodes, edges, notes = load_equity(path)
    if not nodes:
        raise SystemExit('未解析到任何节点，请检查 Excel 表头与数据行')
    apply_special_rules(nodes, notes)
    unify_widths(nodes, mode)
    layout(nodes, edges, MAX_PER_ROW_ORG if mode == 'org' else MAX_PER_ROW_EQUITY)
    wires, labels = build_wires(nodes, edges)

    x0 = min(n['x'] - n['w'] / 2 for n in nodes) - H_UNIT * 0.5
    x1 = max(n['x'] + n['w'] / 2 for n in nodes) + H_UNIT * 0.5
    y0 = min(n['y'] - n['h'] / 2 for n in nodes) - H_UNIT * 0.6
    y1 = max(n['y'] + n['h'] / 2 for n in nodes) + H_UNIT * 0.9   # 底部留折叠徽章空间
    out_nodes = []
    for n in nodes:
        d = {'id': n['id'], 'name': n['name'], 'geo': n.get('geo', ''),
             'type': n['type'], 'x': n['x'], 'y': n['y'],
             'w': round(n['w'], 1), 'h': n['h'],
             'nameFont': n['nameFont'], 'layer': n.get('dlayer', n['layer'])}
        for k in ('face', 'text', 'function', 'leader', 'headcount', 'regular', 'outsourced'):
            if k in n:
                d[k] = n[k]
        out_nodes.append(d)
    return {'viewBox': {'x': round(x0, 1), 'y': round(y0, 1),
                        'w': round(x1 - x0, 1), 'h': round(y1 - y0, 1)},
            'nodes': out_nodes, 'edges': edges, 'wires': wires, 'labels': labels}

def main():
    ap = argparse.ArgumentParser(description='股权/组织架构图 · Excel → 布局 JSON / 独立 HTML')
    ap.add_argument('input', help='输入 Excel（股权模版或组织模版）')
    ap.add_argument('--mode', choices=['auto', 'equity', 'org'], default='auto')
    ap.add_argument('--title', default='', help='图标题（写入 <title>，浏览器标签页显示）')
    ap.add_argument('--json', help='输出 JSON 文件路径')
    ap.add_argument('--html', help='输出独立可交互 HTML 文件路径')
    ap.add_argument('--template', help='HTML 模板路径（默认用 skill 自带模板）')
    args = ap.parse_args()

    mode = args.mode if args.mode != 'auto' else detect_mode(args.input)
    data = build(args.input, mode)
    payload = json.dumps(data, ensure_ascii=False)

    if args.json:
        Path(args.json).write_text(payload, encoding='utf-8')
    if args.html:
        tpl = Path(args.template) if args.template else \
            Path(__file__).resolve().parent.parent / 'assets' / 'widget-template' / 'index.html'
        html = tpl.read_text(encoding='utf-8')
        title = args.title or Path(args.input).stem
        html = html.replace('/*__CHART_TITLE__*/架构图', title)
        html = html.replace('/*__EQUITY_DATA__*/ null', payload, 1)
        Path(args.html).write_text(html, encoding='utf-8')
    if not args.json and not args.html:
        print(payload)
    else:
        print(f'OK mode={mode} nodes={len(data["nodes"])} edges={len(data["edges"])} '
              f'wires={len(data["wires"])} labels={len(data["labels"])}', file=sys.stderr)

if __name__ == '__main__':
    main()
